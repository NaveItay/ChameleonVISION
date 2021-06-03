from openpyxl import Workbook, load_workbook


class Statistics:

    name = "WorldCup"
    teamA = "Israel"
    teamB = "Brazil"
    Location = "Tel Aviv"
    Date = "25/05/2021"
    Weather = "Sunny"

    TeamA_Accuracy_BallIn = 48
    TeamB_Accuracy_BallIn = 52

    TeamA_Accuracy_BallOut = 37
    TeamB_Accuracy_BallOut = 63

    true_alert = 0
    false_alert = 0
    systemAccuracy = 81

    def save(self, true_alert, alerts):

        # Check system accuracy
        self.systemAccuracy = int((true_alert / alerts) * 100)

        self.saveStat(self.name, self.teamA, self.teamB, self.Location, self.Date,
                      self.Weather, self.TeamA_Accuracy_BallIn, self.TeamB_Accuracy_BallIn,
                      self.TeamA_Accuracy_BallOut, self.TeamB_Accuracy_BallOut, self.systemAccuracy)

    def saveStat(self, name, teamA, teamB, Location, Date, Weather,
                 TeamA_Accuracy_BallIn, TeamB_Accuracy_BallIn,
                 TeamA_Accuracy_BallOut, TeamB_Accuracy_BallOut, systemAccuracy):

        wb = load_workbook('/home/chameleonvision/Desktop/FinalProject/ChameleonVISION/statistics/statistics.xlsx')
        ws = wb.active
        ws['A2'].value = name
        ws['B2'].value = teamA
        ws['C2'].value = teamB
        ws['D2'].value = Location
        ws['E2'].value = Date
        ws['F2'].value = Weather
        ws['G2'].value = TeamA_Accuracy_BallIn
        ws['H2'].value = TeamB_Accuracy_BallIn
        ws['I2'].value = TeamA_Accuracy_BallOut
        ws['J2'].value = TeamB_Accuracy_BallOut
        ws['K2'].value = systemAccuracy

        wb.save('/home/chameleonvision/Desktop/FinalProject/ChameleonVISION/statistics/statistics.xlsx')




